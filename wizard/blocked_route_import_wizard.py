import base64
import io
import logging
from datetime import datetime, date
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None
    _logger.warning("La libreria 'openpyxl' no esta instalada. Instale openpyxl para importar archivos Excel.")

class BlockRoutesBlockedImportWizard(models.TransientModel):
    _name = 'block_routes.blocked_import_wizard'
    _description = 'Asistente de Importacion de Bloqueos de Rutas'

    excel_file = fields.Binary(string="Archivo Excel (.xlsx)", required=True)
    filename = fields.Char(string="Nombre de Archivo")

    def action_import_excel(self):
        if not openpyxl:
            raise UserError("La libreria 'openpyxl' no esta instalada en el servidor. Póngase en contacto con su administrador.")

        # Decode base64 file
        try:
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise ValidationError(f"Error al decodificar y leer el archivo Excel: {str(e)}")

        sheet = workbook.active
        Route = self.env['block_routes.route']
        Blocked = self.env['block_routes.blocked']
        
        imported_count = 0
        errors = []

        # Recorremos a partir de la fila 2 para saltar la cabecera
        for r_idx in range(2, sheet.max_row + 1):
            route_name = sheet.cell(row=r_idx, column=1).value
            reason = sheet.cell(row=r_idx, column=2).value
            date_start_val = sheet.cell(row=r_idx, column=3).value
            date_end_val = sheet.cell(row=r_idx, column=4).value

            # Si la fila entera esta vacia, la saltamos
            if not route_name and not reason and not date_start_val and not date_end_val:
                continue

            if not route_name:
                errors.append(f"Fila {r_idx}: El identificador de la ruta está vacío.")
                continue

            route_name = str(route_name).strip()
            
            # Buscar la ruta por nombre
            route_rec = Route.search([('name', '=', route_name)], limit=1)
            if not route_rec:
                errors.append(f"Fila {r_idx}: No se encontró ninguna ruta con el identificador '{route_name}' en Odoo.")
                continue

            # Parsear Fecha de Inicio
            date_start = None
            if isinstance(date_start_val, (datetime, date)):
                date_start = date_start_val
            elif date_start_val:
                try:
                    date_start = fields.Datetime.to_datetime(str(date_start_val).strip())
                except Exception:
                    errors.append(f"Fila {r_idx}: Formato inválido para Fecha de Inicio ('{date_start_val}'). Debe ser AAAA-MM-DD HH:MM:SS.")
                    continue

            # Parsear Fecha de Fin
            date_end = None
            if isinstance(date_end_val, (datetime, date)):
                date_end = date_end_val
            elif date_end_val:
                try:
                    date_end = fields.Datetime.to_datetime(str(date_end_val).strip())
                except Exception:
                    errors.append(f"Fila {r_idx}: Formato inválido para Fecha de Fin ('{date_end_val}'). Debe ser AAAA-MM-DD HH:MM:SS.")
                    continue

            if not date_start or not date_end:
                errors.append(f"Fila {r_idx}: Las fechas de inicio y fin son obligatorias.")
                continue

            if date_end <= date_start:
                errors.append(f"Fila {r_idx}: La fecha de fin es anterior o igual a la de inicio.")
                continue

            # Registrar el bloqueo
            try:
                Blocked.create({
                    'route_id': route_rec.id,
                    'reason': str(reason or 'Bloqueo Importado'),
                    'date_start': date_start,
                    'date_end': date_end
                })
                imported_count += 1
            except Exception as e:
                errors.append(f"Fila {r_idx}: Error al guardar el bloqueo: {str(e)}")

        if errors:
            # Si hay errores, hacemos rollback lanzando una ValidationError detallada
            error_message = "\n".join(errors[:10])
            if len(errors) > 10:
                error_message += f"\n... y {len(errors) - 10} errores más."
            raise ValidationError(f"Se encontraron errores durante la importación. Se cancelaron todos los cambios:\n\n{error_message}")

        # Mensaje de éxito
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Importación Exitosa',
                'message': f'Se importaron exitosamente {imported_count} bloqueos de rutas.',
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'}
            }
        }
